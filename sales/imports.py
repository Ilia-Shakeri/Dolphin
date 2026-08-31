"""Create products in bulk from a filled spreadsheet export.

The operator exports the catalogue, writes new rows on that file, and uploads it
back. Because the header row is one this codebase wrote, columns are matched by
**name** rather than by position — a reordered or partially deleted sheet still
imports correctly, and a sheet that is not ours is refused outright instead of
being read positionally into the wrong fields.

Three outcomes are counted and reported separately, because they mean different
things to the person who uploaded the file:

* **created** — the row became a product;
* **duplicates** — the SKU already exists, so the row was skipped and nothing
  was overwritten. An import never edits an existing product: the operator asked
  to add products, and silently rewriting one they had already priced would be a
  different and much more dangerous operation;
* **invalid** — the row could not be read as a product at all, with the reason.

Everything runs through `create_product`, the same service the form uses, so an
imported product gets the same validation, the same permission check and the
same audit row as one typed by hand.
"""

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook

from common.exceptions import BusinessRuleError
from reports.xlsx import PRODUCT_HEADERS, PRODUCT_UNIT_INPUT
from sales.models import Product, ProductCategory
from sales.services import create_product


#: Columns a row cannot be a product without.
REQUIRED_COLUMNS = ("sku", "name", "current_price")
#: Rows accepted in one upload. A file larger than this is a data-migration
#: task, not a panel action, and should not be attempted through a web request.
MAX_IMPORT_ROWS = 5000


@dataclass
class ImportResult:
    created: int = 0
    duplicates: int = 0
    invalid: int = 0
    #: Bounded so an unreadable file cannot return a megabyte of complaints.
    errors: list = field(default_factory=list)

    def note_error(self, row_number, message):
        self.invalid += 1
        if len(self.errors) < 20:
            self.errors.append({"row": row_number, "detail": message})


def _header_index(sheet):
    """Map our column names to their positions in the uploaded sheet."""
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise BusinessRuleError({"file": "فایل اکسل خالی است."})
    index = {}
    for position, value in enumerate(header_row):
        name = str(value or "").strip().lower()
        if name in PRODUCT_HEADERS:
            index[name] = position
    missing = [name for name in REQUIRED_COLUMNS if name not in index]
    if missing:
        raise BusinessRuleError({
            "file": (
                "ستون‌های "
                f"{', '.join(missing)} در فایل وجود ندارد. ابتدا فهرست کالاها را خروجی بگیرید و روی همان فایل بنویسید."
            )
        })
    return index


def _cell(row, index, name):
    position = index.get(name)
    if position is None or position >= len(row):
        return ""
    value = row[position]
    if value is None:
        return ""
    # openpyxl hands back a float for a numeric cell; str() on that would put a
    # trailing `.0` into a text field.
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _price(raw):
    """A price cell as a Decimal, accepting the grouping the panel displays."""
    text = str(raw).replace("،", "").replace(",", "").strip()
    if not text:
        raise ValueError("قیمت الزامی است.")
    try:
        value = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError("قیمت باید عدد باشد.") from exc
    if value <= 0:
        raise ValueError("قیمت باید بیشتر از صفر باشد.")
    return value


@transaction.atomic
def import_products_from_workbook(*, actor, stream):
    """Read an uploaded workbook and create every product it describes.

    One transaction for the whole file: a run that fails halfway leaves no
    half-imported catalogue behind. Rows that are skipped as duplicate or
    invalid do not fail the run — they are reported.
    """
    try:
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises several unrelated types
        raise BusinessRuleError({"file": "فایل به‌عنوان اکسل قابل خواندن نبود."}) from exc

    sheet = workbook.active
    index = _header_index(sheet)
    result = ImportResult()

    # Existing SKUs are read once. Case-folded, because a spreadsheet round trip
    # routinely changes the case of a code and two products differing only by
    # case are the same product to the person who typed them.
    existing = {sku.casefold() for sku in Product.objects.values_list("sku", flat=True)}
    categories = {
        code.casefold(): pk
        for pk, code in ProductCategory.objects.values_list("pk", "code")
    }
    seen_in_file = set()

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if row_number - 1 > MAX_IMPORT_ROWS:
            raise BusinessRuleError({
                "file": f"هر بارگذاری حداکثر می‌تواند {MAX_IMPORT_ROWS} ردیف داشته باشد."
            })
        if row is None or all(value in (None, "") for value in row):
            continue

        sku = _cell(row, index, "sku")
        name = _cell(row, index, "name")
        if not sku or not name:
            result.note_error(row_number, "درج کد و نام الزامی است.")
            continue

        key = sku.casefold()
        if key in existing or key in seen_in_file:
            result.duplicates += 1
            continue

        try:
            price = _price(_cell(row, index, "current_price"))
        except ValueError as exc:
            result.note_error(row_number, str(exc))
            continue

        unit_text = _cell(row, index, "unit")
        unit = PRODUCT_UNIT_INPUT.get(unit_text, "") if unit_text else ""
        if unit_text and not unit:
            result.note_error(row_number, f"واحد نامعتبر است: {unit_text}")
            continue

        category_code = _cell(row, index, "category_code")
        category = None
        if category_code:
            category_pk = categories.get(category_code.casefold())
            if category_pk is None:
                result.note_error(row_number, f"کد دسته‌بندی نامعتبر است: {category_code}")
                continue
            category = ProductCategory.objects.get(pk=category_pk)

        values = {
            "sku": sku,
            "name": name,
            "current_price": price,
            "brand": _cell(row, index, "brand"),
            "unit": unit,
            "description": _cell(row, index, "description"),
        }
        if category is not None:
            values["category"] = category

        try:
            create_product(actor=actor, **values)
        except BusinessRuleError as exc:
            result.note_error(row_number, str(exc.detail))
            continue

        seen_in_file.add(key)
        result.created += 1

    return result
