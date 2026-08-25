"""Two customer books, and who may read and write each.

The panel hides the حقوقی switch from a marketer. These tests are about what
happens when the hiding is bypassed — because the hiding is not the
authorisation. `customers_for` confines a marketer to the individual book in
the database, and `_validate_customer_kind` refuses them a legal customer on the
way in, so both directions are closed whatever the page rendered.
"""

import io

from django.test import TestCase
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from accounts.models import User
from common.exceptions import BusinessPermissionDenied, BusinessRuleError
from reports.xlsx import CUSTOMER_DIRECTORY_HEADERS
from sales.customer_imports import import_customers_from_workbook
from sales.models import Customer
from sales.selectors import customers_for
from sales.services import create_customer_with_phone, update_customer


PASSWORD = "Strong-pass-937!"


def sheet_from(*rows, headers=CUSTOMER_DIRECTORY_HEADERS):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    stream.name = "customers.xlsx"
    return stream


def row(full_name, *, phone="", national_id="", kind="حقیقی", city=""):
    """A row in CUSTOMER_DIRECTORY_HEADERS order."""
    return (
        "", full_name, kind, phone, national_id, "", "", city, "", "", "", "yes", "", "",
    )


class CustomerKindTests(TestCase):
    def setUp(self):
        self.manager = User.objects.create_user(
            username="kind.manager", password=PASSWORD, role=User.Role.SALES_MANAGER
        )
        self.agent = User.objects.create_user(
            username="kind.agent", password=PASSWORD, role=User.Role.SALES_AGENT
        )

    def client_for(self, user):
        client = APIClient()
        client.force_authenticate(user)
        return client

    def make(self, actor, name, kind, phone=None):
        return create_customer_with_phone(
            actor=actor,
            full_name=name,
            kind=kind,
            phone={"raw_phone": phone, "is_primary": True} if phone else None,
        )

    # --- the default -------------------------------------------------------

    def test_a_customer_created_without_a_kind_is_an_individual(self):
        """Every customer that predates the field must land in a real book.

        A blank default would leave them in neither list, which for the
        marketer means their own customers disappearing from the only book they
        can see.
        """
        customer = create_customer_with_phone(actor=self.manager, full_name="بدون نوع")
        self.assertEqual(customer.kind, Customer.Kind.INDIVIDUAL)
        self.assertIn(customer, customers_for(self.agent.__class__.objects.get(pk=self.manager.pk)))

    # --- read scope --------------------------------------------------------

    def test_a_marketer_reads_only_the_individual_book(self):
        mine_individual = self.make(self.agent, "مشتری حقیقی من", Customer.Kind.INDIVIDUAL)
        mine_legal = Customer.objects.create(
            full_name="شرکت من", kind=Customer.Kind.LEGAL, created_by=self.agent
        )
        theirs = self.make(self.manager, "شرکت دیگری", Customer.Kind.LEGAL)

        visible = set(customers_for(self.agent))
        self.assertEqual(visible, {mine_individual})
        self.assertNotIn(mine_legal, visible)
        self.assertNotIn(theirs, visible)

    def test_a_marketer_asking_for_the_legal_book_gets_an_empty_page(self):
        """Not someone else's customers, and not an error that leaks a count."""
        self.make(self.agent, "حقیقی", Customer.Kind.INDIVIDUAL)
        Customer.objects.create(
            full_name="حقوقی", kind=Customer.Kind.LEGAL, created_by=self.agent
        )
        response = self.client_for(self.agent).get("/api/v1/customers/?kind=legal")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 0)

    def test_a_marketer_cannot_open_a_legal_customer_by_id(self):
        legal = Customer.objects.create(
            full_name="شرکت", kind=Customer.Kind.LEGAL, created_by=self.agent
        )
        response = self.client_for(self.agent).get(f"/api/v1/customers/{legal.pk}/")
        self.assertEqual(response.status_code, 404)

    def test_a_manager_reads_both_books_and_can_narrow_to_either(self):
        self.make(self.manager, "شخص", Customer.Kind.INDIVIDUAL)
        self.make(self.manager, "شرکت", Customer.Kind.LEGAL)
        client = self.client_for(self.manager)
        self.assertEqual(client.get("/api/v1/customers/").data["count"], 2)
        self.assertEqual(client.get("/api/v1/customers/?kind=individual").data["count"], 1)
        self.assertEqual(client.get("/api/v1/customers/?kind=legal").data["count"], 1)

    def test_an_unknown_kind_is_a_request_error_not_a_silent_full_list(self):
        self.make(self.manager, "شخص", Customer.Kind.INDIVIDUAL)
        response = self.client_for(self.manager).get("/api/v1/customers/?kind=corporation")
        self.assertEqual(response.status_code, 400)

    # --- write scope -------------------------------------------------------

    def test_a_marketer_cannot_create_a_legal_customer(self):
        with self.assertRaises(BusinessPermissionDenied):
            create_customer_with_phone(
                actor=self.agent, full_name="شرکت", kind=Customer.Kind.LEGAL
            )
        self.assertEqual(Customer.objects.count(), 0)

    def test_the_api_refuses_a_marketer_a_legal_customer(self):
        response = self.client_for(self.agent).post(
            "/api/v1/customers/", {"full_name": "شرکت", "kind": "legal"}, format="json"
        )
        self.assertEqual(response.status_code, 403)
        self.assertEqual(Customer.objects.count(), 0)

    def test_a_marketer_cannot_move_their_own_customer_into_the_legal_book(self):
        mine = self.make(self.agent, "شخص", Customer.Kind.INDIVIDUAL)
        with self.assertRaises(BusinessPermissionDenied):
            update_customer(actor=self.agent, customer=mine, kind=Customer.Kind.LEGAL)
        mine.refresh_from_db()
        self.assertEqual(mine.kind, Customer.Kind.INDIVIDUAL)

    def test_an_unknown_kind_is_refused_by_name_rather_than_by_the_constraint(self):
        """The database would catch it too, but as an error naming no field."""
        with self.assertRaises(BusinessRuleError) as caught:
            create_customer_with_phone(
                actor=self.manager, full_name="نامعتبر", kind="corporation"
            )
        self.assertIn("kind", caught.exception.detail)

    def test_a_manager_may_create_in_either_book(self):
        for kind in (Customer.Kind.INDIVIDUAL, Customer.Kind.LEGAL):
            with self.subTest(kind=kind):
                customer = self.make(self.manager, f"مشتری {kind}", kind)
                self.assertEqual(customer.kind, kind)


class CustomerSpreadsheetTests(TestCase):
    """Export one book, write on the file, upload it back into a book."""

    def setUp(self):
        self.manager = User.objects.create_user(
            username="sheet.manager", password=PASSWORD, role=User.Role.SALES_MANAGER
        )
        self.agent = User.objects.create_user(
            username="sheet.agent", password=PASSWORD, role=User.Role.SALES_AGENT
        )

    def client_for(self, user):
        client = APIClient()
        client.force_authenticate(user)
        return client

    def test_an_export_is_a_file_the_import_understands(self):
        create_customer_with_phone(
            actor=self.manager,
            full_name="شرکت الف",
            kind=Customer.Kind.LEGAL,
            city="تهران",
            phone={"raw_phone": "09121110001", "is_primary": True},
        )
        response = self.client_for(self.manager).get("/api/v1/exports/customers.xlsx?kind=legal")
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        sheet.append(row("شرکت ب", phone="09121110002", kind="حقوقی", city="کرج"))
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)

        result = import_customers_from_workbook(
            actor=self.manager, stream=stream, kind=Customer.Kind.LEGAL
        )
        # The exported row returns as a duplicate; only the written one is new.
        self.assertEqual(result.created, 1)
        self.assertEqual(result.duplicates, 1)
        self.assertEqual(result.invalid, 0)
        created = Customer.objects.get(full_name="شرکت ب")
        self.assertEqual(created.kind, Customer.Kind.LEGAL)
        self.assertEqual(created.city, "کرج")

    def test_the_export_is_narrowed_to_the_chosen_book(self):
        create_customer_with_phone(
            actor=self.manager, full_name="شخص", kind=Customer.Kind.INDIVIDUAL
        )
        create_customer_with_phone(
            actor=self.manager, full_name="شرکت", kind=Customer.Kind.LEGAL
        )
        client = self.client_for(self.manager)
        for kind, expected in (("individual", {"شخص"}), ("legal", {"شرکت"})):
            with self.subTest(kind=kind):
                payload = client.get(f"/api/v1/exports/customers.xlsx?kind={kind}").content
                sheet = load_workbook(io.BytesIO(payload))["customers"]
                names = {
                    line[1]
                    for line in sheet.iter_rows(min_row=2, values_only=True)
                    if line[0] is not None
                }
                self.assertEqual(names, expected)

    def test_a_marketer_exports_only_their_own_individual_book(self):
        create_customer_with_phone(
            actor=self.agent, full_name="مال من", kind=Customer.Kind.INDIVIDUAL
        )
        Customer.objects.create(
            full_name="شرکت پنهان", kind=Customer.Kind.LEGAL, created_by=self.agent
        )
        payload = self.client_for(self.agent).get(
            "/api/v1/exports/customers.xlsx?kind=legal"
        ).content
        sheet = load_workbook(io.BytesIO(payload))["customers"]
        names = {
            line[1] for line in sheet.iter_rows(min_row=2, values_only=True)
            if line[0] is not None
        }
        self.assertEqual(names, set())

    def test_the_chosen_list_wins_over_the_kind_written_in_the_file(self):
        """The operator picked a list to import into; a stray cell is not that."""
        result = import_customers_from_workbook(
            actor=self.manager,
            stream=sheet_from(row("شرکت", phone="09121110003", kind="حقیقی")),
            kind=Customer.Kind.LEGAL,
        )
        self.assertEqual(result.created, 1)
        self.assertEqual(Customer.objects.get(full_name="شرکت").kind, Customer.Kind.LEGAL)

    def test_a_repeated_phone_is_a_duplicate_and_overwrites_nothing(self):
        create_customer_with_phone(
            actor=self.manager,
            full_name="نام اصلی",
            phone={"raw_phone": "09121110004", "is_primary": True},
        )
        result = import_customers_from_workbook(
            actor=self.manager,
            stream=sheet_from(row("نام تازه", phone="09121110004")),
            kind=Customer.Kind.INDIVIDUAL,
        )
        self.assertEqual(result.created, 0)
        self.assertEqual(result.duplicates, 1)
        self.assertEqual(Customer.objects.get(full_name="نام اصلی").full_name, "نام اصلی")

    def test_the_same_phone_written_two_ways_is_still_one_person(self):
        """`+989121110005` and `09121110005` are the same number."""
        result = import_customers_from_workbook(
            actor=self.manager,
            stream=sheet_from(
                row("یک", phone="09121110005"),
                row("دو", phone="+989121110005"),
            ),
            kind=Customer.Kind.INDIVIDUAL,
        )
        self.assertEqual(result.created, 1)
        self.assertEqual(result.duplicates, 1)

    def test_a_repeated_national_id_is_a_duplicate_even_without_a_phone(self):
        create_customer_with_phone(
            actor=self.manager, full_name="اصلی", national_id="0012345678"
        )
        result = import_customers_from_workbook(
            actor=self.manager,
            stream=sheet_from(row("تکراری", national_id="0012345678")),
            kind=Customer.Kind.INDIVIDUAL,
        )
        self.assertEqual(result.created, 0)
        self.assertEqual(result.duplicates, 1)

    def test_a_row_without_a_name_is_invalid_and_the_rest_still_import(self):
        result = import_customers_from_workbook(
            actor=self.manager,
            stream=sheet_from(
                row("خوب", phone="09121110006"),
                row("", phone="09121110007"),
                row("شماره خراب", phone="نه-یک-شماره"),
                row("خوب دوم", phone="09121110008"),
            ),
            kind=Customer.Kind.INDIVIDUAL,
        )
        self.assertEqual(result.created, 2)
        self.assertEqual(result.invalid, 2)
        self.assertEqual([entry["row"] for entry in result.errors], [3, 4])

    def test_a_sheet_that_is_not_ours_is_refused(self):
        with self.assertRaises(BusinessRuleError):
            import_customers_from_workbook(
                actor=self.manager,
                stream=sheet_from(("a", "b"), headers=("alpha", "beta")),
                kind=Customer.Kind.INDIVIDUAL,
            )
        self.assertEqual(Customer.objects.count(), 0)

    def test_an_unknown_target_list_is_refused_before_anything_is_read(self):
        with self.assertRaises(BusinessRuleError):
            import_customers_from_workbook(
                actor=self.manager,
                stream=sheet_from(row("کسی", phone="09121110009")),
                kind="corporation",
            )
        self.assertEqual(Customer.objects.count(), 0)

    def test_a_marketer_cannot_import_into_the_legal_book(self):
        response = self.client_for(self.agent).post(
            "/api/v1/customers/import-xlsx/",
            {"file": sheet_from(row("شرکت", phone="09121110010")), "kind": "legal"},
            format="multipart",
        )
        self.assertEqual(response.status_code, 403)
        self.assertEqual(Customer.objects.count(), 0)

    def test_the_endpoint_reports_all_three_counts(self):
        create_customer_with_phone(
            actor=self.manager,
            full_name="موجود",
            phone={"raw_phone": "09121110011", "is_primary": True},
        )
        upload = sheet_from(
            row("تکراری", phone="09121110011"),
            row("تازه", phone="09121110012"),
            row("", phone="09121110013"),
        )
        response = self.client_for(self.manager).post(
            "/api/v1/customers/import-xlsx/",
            {"file": upload, "kind": "individual"},
            format="multipart",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["created"], 1)
        self.assertEqual(response.data["duplicates"], 1)
        self.assertEqual(response.data["invalid"], 1)

    def test_the_endpoint_refuses_anything_that_is_not_an_xlsx(self):
        response = self.client_for(self.manager).post(
            "/api/v1/customers/import-xlsx/",
            {"file": io.BytesIO(b"full_name\nsomeone"), "kind": "individual"},
            format="multipart",
        )
        self.assertEqual(response.status_code, 400)
