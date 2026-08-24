"""The product catalogue round trip: export, write on the file, upload it back.

The whole design rests on one claim — that the file the export produces is a
file the import understands. The first test here proves exactly that, by
round-tripping without ever hand-building a sheet. The rest pin the three
outcomes an operator is told about: created, duplicate, invalid.
"""

import io
from decimal import Decimal

from django.test import TestCase
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from accounts.models import User
from common.exceptions import BusinessRuleError
from reports.xlsx import PRODUCT_HEADERS
from sales.imports import import_products_from_workbook
from sales.models import Product
from sales.services import create_product, create_product_category


PASSWORD = "Strong-pass-937!"


def sheet_from(*rows, headers=PRODUCT_HEADERS):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def row(sku, name, price, *, category="", brand="", unit="", description=""):
    """A row in PRODUCT_HEADERS order, with `id` blank as an import ignores it."""
    return ("", sku, name, category, brand, unit, price, description, "yes")


class ProductSpreadsheetTests(TestCase):
    def setUp(self):
        self.manager = User.objects.create_user(
            username="cat.manager", password=PASSWORD, role=User.Role.SALES_MANAGER
        )
        self.agent = User.objects.create_user(
            username="cat.agent", password=PASSWORD, role=User.Role.SALES_AGENT
        )
        self.category = create_product_category(
            actor=self.manager, code="STA", name="لوازم‌التحریر"
        )

    def client_for(self, user):
        client = APIClient()
        client.force_authenticate(user)
        return client

    # --- the round trip ----------------------------------------------------

    def test_an_export_is_a_file_the_import_understands(self):
        """The claim the whole feature rests on, proved end to end."""
        create_product(
            actor=self.manager,
            sku="RT-1",
            name="دفتر",
            category=self.category,
            brand="آرشیو",
            unit=Product.Unit.CARTON,
            current_price=Decimal("125000.00"),
        )
        response = self.client_for(self.manager).get("/api/v1/exports/products.xlsx")
        self.assertEqual(response.status_code, 200)

        # Write one new row onto the exported file, exactly as an operator would.
        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        sheet.append(("", "RT-2", "خودکار", "STA", "آرشیو", "جعبه", 4500, "", "yes"))
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)

        result = import_products_from_workbook(actor=self.manager, stream=stream)

        # The exported row comes back as a duplicate — never a second product,
        # and never an overwrite of the first.
        self.assertEqual(result.created, 1)
        self.assertEqual(result.duplicates, 1)
        self.assertEqual(result.invalid, 0)
        self.assertEqual(Product.objects.filter(sku="RT-1").count(), 1)

        created = Product.objects.get(sku="RT-2")
        self.assertEqual(created.name, "خودکار")
        self.assertEqual(created.category, self.category)
        self.assertEqual(created.unit, Product.Unit.BOX)
        self.assertEqual(created.current_price, Decimal("4500"))

    def test_the_export_carries_the_unit_as_its_persian_label(self):
        create_product(
            actor=self.manager,
            sku="U-1",
            name="برنج",
            unit=Product.Unit.KILOGRAM,
            current_price=Decimal("900000.00"),
        )
        response = self.client_for(self.manager).get("/api/v1/exports/products.xlsx")
        sheet = load_workbook(io.BytesIO(response.content)).active
        header, data = list(sheet.iter_rows(values_only=True))[:2]
        self.assertEqual(header, PRODUCT_HEADERS)
        self.assertEqual(data[header.index("unit")], "کیلوگرم")

    # --- duplicates --------------------------------------------------------

    def test_a_duplicate_sku_is_counted_and_never_overwrites(self):
        create_product(
            actor=self.manager, sku="DUP-1", name="نام اصلی", current_price=Decimal("1000.00")
        )
        result = import_products_from_workbook(
            actor=self.manager,
            stream=sheet_from(row("DUP-1", "نام تازه", 9999)),
        )
        self.assertEqual(result.created, 0)
        self.assertEqual(result.duplicates, 1)
        existing = Product.objects.get(sku="DUP-1")
        self.assertEqual(existing.name, "نام اصلی")
        self.assertEqual(existing.current_price, Decimal("1000.00"))

    def test_a_duplicate_is_the_same_code_in_a_different_case(self):
        """A spreadsheet round trip routinely changes the case of a code."""
        create_product(
            actor=self.manager, sku="Case-1", name="کالا", current_price=Decimal("1000.00")
        )
        result = import_products_from_workbook(
            actor=self.manager, stream=sheet_from(row("CASE-1", "کالا", 1000))
        )
        self.assertEqual(result.duplicates, 1)
        self.assertEqual(Product.objects.filter(sku__iexact="case-1").count(), 1)

    def test_a_code_repeated_inside_one_file_is_imported_once(self):
        result = import_products_from_workbook(
            actor=self.manager,
            stream=sheet_from(row("IN-1", "یک", 500), row("IN-1", "دو", 700)),
        )
        self.assertEqual(result.created, 1)
        self.assertEqual(result.duplicates, 1)
        self.assertEqual(Product.objects.get(sku="IN-1").name, "یک")

    # --- rows that cannot become products ----------------------------------

    def test_an_invalid_row_is_reported_and_the_rest_still_import(self):
        result = import_products_from_workbook(
            actor=self.manager,
            stream=sheet_from(
                row("OK-1", "خوب", 1000),
                row("", "بدون کد", 1000),
                row("BAD-2", "قیمت خراب", "abc"),
                row("BAD-3", "قیمت صفر", 0),
                row("BAD-4", "واحد ناشناخته", 1000, unit="بشکه"),
                row("BAD-5", "دستهٔ ناشناخته", 1000, category="NOPE"),
                row("OK-2", "خوب دوم", 2000),
            ),
        )
        self.assertEqual(result.created, 2)
        self.assertEqual(result.invalid, 5)
        self.assertEqual(
            sorted(Product.objects.values_list("sku", flat=True)), ["OK-1", "OK-2"]
        )
        # Each complaint names the row it came from, so the operator can find it.
        self.assertEqual([entry["row"] for entry in result.errors], [3, 4, 5, 6, 7])

    def test_a_blank_row_is_skipped_without_being_called_invalid(self):
        result = import_products_from_workbook(
            actor=self.manager,
            stream=sheet_from(row("B-1", "کالا", 100), (None,) * 9, row("B-2", "کالا", 200)),
        )
        self.assertEqual(result.created, 2)
        self.assertEqual(result.invalid, 0)

    def test_a_price_written_with_the_panel_grouping_is_accepted(self):
        """The operator sees `12،500،000` on screen and may type it back."""
        result = import_products_from_workbook(
            actor=self.manager, stream=sheet_from(row("G-1", "کالا", "12،500،000"))
        )
        self.assertEqual(result.created, 1)
        self.assertEqual(Product.objects.get(sku="G-1").current_price, Decimal("12500000"))

    def test_a_sheet_that_is_not_ours_is_refused_rather_than_read_positionally(self):
        with self.assertRaises(BusinessRuleError):
            import_products_from_workbook(
                actor=self.manager,
                stream=sheet_from(("x", "y", "z"), headers=("alpha", "beta", "gamma")),
            )
        self.assertEqual(Product.objects.count(), 0)

    def test_columns_may_be_reordered_because_they_match_by_name(self):
        headers = ("name", "current_price", "sku")
        result = import_products_from_workbook(
            actor=self.manager,
            stream=sheet_from(("کالای جابه‌جا", 750, "MOVED-1"), headers=headers),
        )
        self.assertEqual(result.created, 1)
        self.assertEqual(Product.objects.get(sku="MOVED-1").name, "کالای جابه‌جا")

    def test_a_file_that_is_not_a_spreadsheet_is_refused(self):
        with self.assertRaises(BusinessRuleError):
            import_products_from_workbook(
                actor=self.manager, stream=io.BytesIO(b"not a workbook at all")
            )

    # --- authorisation -----------------------------------------------------

    def test_an_agent_cannot_import_products(self):
        """Frontend hiding is not authorisation; the endpoint refuses too."""
        upload = sheet_from(row("AG-1", "کالا", 100))
        upload.name = "catalogue.xlsx"
        response = self.client_for(self.agent).post(
            "/api/v1/products/import-xlsx/", {"file": upload}, format="multipart"
        )
        self.assertEqual(response.status_code, 403)
        self.assertEqual(Product.objects.count(), 0)

    def test_the_endpoint_refuses_anything_that_is_not_an_xlsx(self):
        response = self.client_for(self.manager).post(
            "/api/v1/products/import-xlsx/",
            {"file": io.BytesIO(b"id,sku\n1,X")},
            format="multipart",
        )
        self.assertEqual(response.status_code, 400)

    def test_the_endpoint_reports_all_three_counts(self):
        create_product(
            actor=self.manager, sku="EP-1", name="موجود", current_price=Decimal("100.00")
        )
        upload = sheet_from(
            row("EP-1", "تکراری", 100),
            row("EP-2", "تازه", 200),
            row("EP-3", "خراب", "abc"),
        )
        upload.name = "catalogue.xlsx"
        response = self.client_for(self.manager).post(
            "/api/v1/products/import-xlsx/", {"file": upload}, format="multipart"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["created"], 1)
        self.assertEqual(response.data["duplicates"], 1)
        self.assertEqual(response.data["invalid"], 1)

    def test_an_agent_may_still_export_what_it_can_already_read(self):
        create_product(
            actor=self.manager, sku="EX-1", name="کالا", current_price=Decimal("100.00")
        )
        response = self.client_for(self.agent).get("/api/v1/exports/products.xlsx")
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])
