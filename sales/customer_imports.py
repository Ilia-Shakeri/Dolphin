"""Create customers in bulk from a filled spreadsheet export.

The same round trip as `sales.imports` does for products, and deliberately the
same shape: the operator exports a list, writes rows on that file, and uploads
it back, so columns are matched by **header name** rather than by position.

What differs is what makes two rows the same customer. A product has a SKU. A
person does not, so this module uses the identity the rest of the codebase
already uses for a customer — **the phone number**. `CustomerPhone` carries a
database-level uniqueness constraint on the normalized phone among active rows,
which means a duplicated phone is not merely untidy, it is a write the database
would refuse. National ID is checked as well when a row carries one, because two
records for the same national ID are the same person however they were typed.

A row with neither a phone nor a national ID cannot be recognised as a duplicate
of anything, and is created. That is a real limitation and not a bug to hide:
there is nothing in such a row that identifies a person, so no honest comparison
exists.

Every row goes through `create_customer_with_phone`, the service the form uses,
so an imported customer gets the same validation, the same scope check and the
same audit row as one typed by hand — including the rule that a marketer may not
create a legal customer.
"""

from dataclasses import dataclass, field

from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import IntegrityError, transaction
from openpyxl import load_workbook

from common.exceptions import BusinessRuleError
from common.phones import normalize_customer_phone
from reports.xlsx import CUSTOMER_DIRECTORY_HEADERS
from sales.models import Customer, CustomerPhone
from sales.services import create_customer_with_phone


#: A row is not a customer without a name.
REQUIRED_COLUMNS = ("full_name",)
#: Text columns an import may set, in the panel's own field names.
TEXT_COLUMNS = (
    "national_id", "email", "province", "city", "postal_code", "category", "address",
)
MAX_IMPORT_ROWS = 5000


@dataclass
class CustomerImportResult:
    created: int = 0
    duplicates: int = 0
    invalid: int = 0
    errors: list = field(default_factory=list)

    def note_error(self, row_number, message):
        self.invalid += 1
        if len(self.errors) < 20:
            self.errors.append({"row": row_number, "detail": message})


def _header_index(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise BusinessRuleError({"file": "فایل اکسل خالی است."})
    index = {}
    for position, value in enumerate(header_row):
        name = str(value or "").strip().lower()
        if name in CUSTOMER_DIRECTORY_HEADERS:
            index[name] = position
    missing = [name for name in REQUIRED_COLUMNS if name not in index]
    if missing:
        raise BusinessRuleError({
            "file": (
                f"ستون‌های {', '.join(missing)} در فایل وجود ندارد. "
                "ابتدا فهرست را خروجی بگیرید و روی همان فایل بنویسید."
            )
        })
    return index


def _cell(row, index, name):
    position = index.get(name)
    if position is None or position >= len(row):
        return ""
    value = row[position]
    if value is None:
        return ""
    # openpyxl reads a phone typed as digits as a float; str() on that would
    # produce `9121234567.0`.
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@transaction.atomic
def import_customers_from_workbook(*, actor, stream, kind):
    """Read an uploaded workbook and create every customer it describes.

    `kind` is the list the operator chose in the dialog, and it wins over
    anything written in the file's own `kind` column. The operator picked a list
    to import *into*; silently scattering their rows across both books because
    of a stray cell would not be what they asked for.
    """
    if kind not in Customer.Kind.values:
        raise BusinessRuleError({"kind": "نوع مشتری را از فهرست انتخاب کنید."})

    try:
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises several unrelated types
        raise BusinessRuleError({"file": "فایل به‌عنوان اکسل قابل خواندن نبود."}) from exc

    sheet = workbook.active
    index = _header_index(sheet)
    result = CustomerImportResult()

    # Read the existing identities once. Phones are compared in their normalized
    # form, which is the form the uniqueness constraint is written over, so this
    # check and the database agree about what a duplicate is.
    existing_phones = set(
        CustomerPhone.objects.filter(is_active=True).values_list("normalized_phone", flat=True)
    )
    existing_ids = {
        value.casefold()
        for value in Customer.objects.exclude(national_id="").values_list("national_id", flat=True)
    }
    seen_phones = set()
    seen_ids = set()

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row_number - 1 > MAX_IMPORT_ROWS:
            raise BusinessRuleError({"file": f"هر بارگذاری حداکثر می‌تواند {MAX_IMPORT_ROWS} ردیف داشته باشد."})
        if row is None or all(value in (None, "") for value in row):
            continue

        full_name = _cell(row, index, "full_name")
        if not full_name:
            result.note_error(row_number, "نام کامل الزامی است.")
            continue

        raw_phone = _cell(row, index, "primary_phone")
        normalized = ""
        if raw_phone:
            try:
                normalized = normalize_customer_phone(raw_phone)
            except DjangoValidationError:
                result.note_error(row_number, f"شماره تلفن ایرانی معتبر نیست: {raw_phone}")
                continue

        national_id = _cell(row, index, "national_id")
        id_key = national_id.casefold() if national_id else ""

        if normalized and (normalized in existing_phones or normalized in seen_phones):
            result.duplicates += 1
            continue
        if id_key and (id_key in existing_ids or id_key in seen_ids):
            result.duplicates += 1
            continue

        values = {"full_name": full_name, "kind": kind}
        for column in TEXT_COLUMNS:
            values[column] = _cell(row, index, column)

        phone = None
        if raw_phone:
            phone = {"raw_phone": raw_phone, "label": "", "is_primary": True}

        try:
            create_customer_with_phone(actor=actor, phone=phone, **values)
        except BusinessRuleError as exc:
            result.note_error(row_number, str(exc.detail))
            continue
        except IntegrityError as exc:
            # A phone this run did not know about — another session wrote it
            # between the read above and now. It is still a duplicate.
            raise BusinessRuleError({
                "file": f"ردیف {row_number} با مشتری‌ای که هنگام اجرای این بارگذاری ایجاد شد تداخل دارد.",
            }) from exc

        if normalized:
            seen_phones.add(normalized)
        if id_key:
            seen_ids.add(id_key)
        result.created += 1

    return result
