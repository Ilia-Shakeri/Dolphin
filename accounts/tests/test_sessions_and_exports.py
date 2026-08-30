"""Session administration (1.8) and the directory exports (1.9, 2.6)."""

from decimal import Decimal

from django.contrib.sessions.models import Session
from django.test import TestCase

from accounts.models import User
from accounts.sessions import active_sessions_for, revoke_sessions
from auditlog.models import ActivityLog
from common.exceptions import BusinessPermissionDenied
from sales.services import create_customer_with_phone


PASSWORD = "Strong-pass-937!"


class SessionAdministrationTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="sess.admin", password=PASSWORD, role=User.Role.PLATFORM_ADMIN
        )
        self.agent = User.objects.create_user(
            username="sess.agent", password=PASSWORD, role=User.Role.SALES_AGENT
        )

    def sign_in_agent(self):
        """A real login, so the row under test is one the site actually issued."""
        client = self.client_class()
        self.assertTrue(client.login(username=self.agent.username, password=PASSWORD))
        return client

    def test_a_signed_in_user_has_a_listed_session(self):
        self.sign_in_agent()
        rows = active_sessions_for(actor=self.admin, target=self.agent)
        self.assertEqual(len(rows), 1)
        self.assertIn("reference", rows[0])
        self.assertIn("expires_at", rows[0])
        # The session key is the credential itself and is never handed out.
        self.assertNotIn("session_key", rows[0])

    def test_one_users_session_is_not_listed_under_another(self):
        self.sign_in_agent()
        self.assertEqual(active_sessions_for(actor=self.admin, target=self.admin), [])

    def test_revoking_signs_the_browser_out_for_real(self):
        client = self.sign_in_agent()
        self.assertEqual(client.get("/api/v1/auth/me/").status_code, 200)

        ended = revoke_sessions(actor=self.admin, target=self.agent)
        self.assertEqual(ended, 1)
        # The next request from that same browser is no longer authenticated.
        self.assertIn(client.get("/api/v1/auth/me/").status_code, (401, 403))
        self.assertEqual(active_sessions_for(actor=self.admin, target=self.agent), [])

    def test_revoking_does_not_disable_the_account(self):
        self.sign_in_agent()
        revoke_sessions(actor=self.admin, target=self.agent)
        self.agent.refresh_from_db()
        self.assertTrue(self.agent.is_active)
        # And the user can sign straight back in.
        client = self.client_class()
        self.assertTrue(client.login(username=self.agent.username, password=PASSWORD))

    def test_revoking_is_audited_without_recording_the_session_key(self):
        self.sign_in_agent()
        keys = list(Session.objects.values_list("session_key", flat=True))
        references = [
            row["reference"] for row in active_sessions_for(actor=self.admin, target=self.agent)
        ]
        revoke_sessions(actor=self.admin, target=self.agent)
        entry = ActivityLog.objects.filter(operation="user.sessions_revoked").latest("id")
        self.assertEqual(entry.actor, self.admin)
        serialized = str(entry.safe_changes)
        for key in keys:
            self.assertNotIn(key, serialized)
        for reference in references:
            self.assertNotIn(reference, serialized)

    def test_only_a_user_administrator_may_look_or_revoke(self):
        for role in (User.Role.SALES_AGENT, User.Role.SALES_MANAGER, User.Role.COMPANY_IT):
            actor = User.objects.create_user(
                username=f"sess.{role}", password=PASSWORD, role=role
            )
            with self.subTest(role=role):
                with self.assertRaises(BusinessPermissionDenied):
                    active_sessions_for(actor=actor, target=self.agent)
                with self.assertRaises(BusinessPermissionDenied):
                    revoke_sessions(actor=actor, target=self.agent)

    def test_the_api_matches_the_service(self):
        session_key = self.sign_in_agent().session.session_key
        self.client.force_login(self.admin)
        listing = self.client.get(f"/api/v1/users/{self.agent.pk}/sessions/")
        self.assertEqual(listing.status_code, 200)
        payload = listing.json()
        self.assertEqual(payload["count"], 1)
        # The response carries a reference, never the key that would let its
        # reader sign in as that user.
        self.assertNotIn("session_key", listing.content.decode("utf-8"))
        self.assertNotIn(session_key, listing.content.decode("utf-8"))
        self.assertTrue(payload["results"][0]["reference"])

        revoked = self.client.post(
            f"/api/v1/users/{self.agent.pk}/revoke-sessions/",
            data={},
            content_type="application/json",
        )
        self.assertEqual(revoked.status_code, 200)
        self.assertEqual(revoked.json()["ended"], 1)

    def test_a_non_administrator_is_refused_by_the_api(self):
        self.client.force_login(self.agent)
        self.assertEqual(
            self.client.get(f"/api/v1/users/{self.agent.pk}/sessions/").status_code, 403
        )

    def test_an_expired_session_is_neither_listed_nor_counted(self):
        from django.utils import timezone

        self.sign_in_agent()
        Session.objects.update(expire_date=timezone.now() - timezone.timedelta(minutes=1))
        self.assertEqual(active_sessions_for(actor=self.admin, target=self.agent), [])
        self.assertEqual(revoke_sessions(actor=self.admin, target=self.agent), 0)


class DirectoryExportTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username="dir.admin", password=PASSWORD, role=User.Role.PLATFORM_ADMIN
        )
        self.agent = User.objects.create_user(
            username="dir.agent", password=PASSWORD, role=User.Role.SALES_AGENT
        )
        self.other = User.objects.create_user(
            username="dir.other", password=PASSWORD, role=User.Role.SALES_AGENT
        )
        self.mine = create_customer_with_phone(
            actor=self.agent,
            full_name="مشتری خودم",
            phone={"raw_phone": "09121110001", "is_primary": True},
        )
        self.theirs = create_customer_with_phone(
            actor=self.other,
            full_name="مشتری دیگری",
            phone={"raw_phone": "09121110002", "is_primary": True},
        )

    def test_the_user_export_is_platform_admin_only(self):
        self.client.force_login(self.admin)
        response = self.client.get("/api/v1/exports/users.xlsx")
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])
        self.assertIn("dolphin-users.xlsx", response["Content-Disposition"])

        for role_user in (self.agent, self.other):
            self.client.force_login(role_user)
            self.assertEqual(self.client.get("/api/v1/exports/users.xlsx").status_code, 403)

    def test_the_user_export_carries_no_credential_material(self):
        self.client.force_login(self.admin)
        payload = self.client.get("/api/v1/exports/users.xlsx").content
        for secret in (self.admin.password.encode(), b"pbkdf2", b"session"):
            self.assertNotIn(secret, payload)

    def test_the_customer_export_is_scoped_exactly_like_the_list(self):
        self.client.force_login(self.agent)
        response = self.client.get("/api/v1/exports/customers.xlsx")
        self.assertEqual(response.status_code, 200)
        self.assertIn("dolphin-customers.xlsx", response["Content-Disposition"])

        listed = self.client.get("/api/v1/customers/").json()
        names = {row["full_name"] for row in listed["results"]}
        self.assertIn("مشتری خودم", names)
        self.assertNotIn("مشتری دیگری", names)

        # The workbook shows the same set: exactly one data row for this agent.
        from io import BytesIO

        from openpyxl import load_workbook

        sheet = load_workbook(BytesIO(response.content))["customers"]
        rows = [row for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][1], "مشتری خودم")

    def test_a_manager_sees_every_customer_in_the_same_export(self):
        manager = User.objects.create_user(
            username="dir.manager", password=PASSWORD, role=User.Role.SALES_MANAGER
        )
        self.client.force_login(manager)
        from io import BytesIO

        from openpyxl import load_workbook

        payload = self.client.get("/api/v1/exports/customers.xlsx").content
        sheet = load_workbook(BytesIO(payload))["customers"]
        names = {row[1] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None}
        self.assertEqual(names, {"مشتری خودم", "مشتری دیگری"})

    def test_the_customer_export_is_absent_when_the_feature_is_disabled(self):
        from common.deployment.profile import DeploymentProfile, override_active_profile
        from common.deployment.registry import ALL_FEATURES

        reduced = DeploymentProfile(
            profile_id="client-1",
            features=frozenset(ALL_FEATURES) - {"customers"},
            source="signed-manifest",
        )
        self.client.force_login(self.admin)
        with override_active_profile(reduced):
            self.assertEqual(self.client.get("/api/v1/exports/customers.xlsx").status_code, 404)
