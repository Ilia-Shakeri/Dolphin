from io import BytesIO

from openpyxl import Workbook

from common import jalali
from reports.services import UserPerformanceReport


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
REPORT_HEADERS = (
    "user_id",
    "username",
    "customers_created_count",
    "sales_count",
    "sales_amount",
    "average_sale_amount",
)
FORMULA_PREFIXES = ("=", "+", "-", "@")


def spreadsheet_date(value):
    """A stored Gregorian value as the Jalali text a Client-1 user expects.

    `BIZ-007`: the export is client-facing, so its dates are Jalali. The column
    names stay machine identifiers and the underlying record is untouched — this
    converts the displayed value only, and claims no accounting meaning.

    Written as text rather than a date cell on purpose: a spreadsheet has no
    Jalali date type, so a real date cell would be silently re-rendered as
    Gregorian by the reader's locale.
    """
    return jalali.format_date(value) if value not in (None, "") else ""


def spreadsheet_datetime(value):
    """As `spreadsheet_date`, keeping the Tehran-local time of day."""
    return jalali.format_datetime(value) if value not in (None, "") else ""


def safe_spreadsheet_text(value: str) -> str:
    if value and (
        value != value.lstrip(" \t\r\n")
        or value.startswith(FORMULA_PREFIXES)
    ):
        return f"'{value}"
    return value


def build_user_performance_workbook(report: UserPerformanceReport) -> bytes:
    workbook = Workbook()
    workbook.iso_dates = True
    workbook.properties.creator = "ForooshBin"

    sheet = workbook.active
    sheet.title = "user-performance"
    sheet.freeze_panes = "A2"
    sheet.append(REPORT_HEADERS)
    for row in report.results:
        sheet.append(
            (
                row.user_id,
                safe_spreadsheet_text(row.username),
                row.customers_created_count,
                row.sales_count,
                format(row.sales_amount, ".2f"),
                format(row.average_sale_amount, ".2f"),
            )
        )
    if sheet.max_row > 1:
        for cells in sheet.iter_rows(min_row=2, min_col=5, max_col=6):
            for cell in cells:
                cell.number_format = "@"
    sheet.auto_filter.ref = sheet.dimensions

    summary = workbook.create_sheet("summary")
    summary.append(("metric", "value"))
    summary.append(("customers_created_count", report.summary.customers_created_count))
    summary.append(("sales_count", report.summary.sales_count))
    summary.append(("sales_amount", format(report.summary.sales_amount, ".2f")))
    summary.append(("average_sale_amount", format(report.summary.average_sale_amount, ".2f")))
    summary["B4"].number_format = "@"
    summary["B5"].number_format = "@"

    filters = workbook.create_sheet("filters")
    filters.append(("field", "value"))
    # The `filters` sheet is the normalized query echoed back, and a documented
    # contract holds it identical to the JSON response — so the canonical ISO
    # value stays, and the Jalali rendering is added beside it rather than
    # replacing it. The summary sheet and the data columns, which are read
    # rather than compared, show Jalali alone.
    filters.append(("period_start", report.period_start))
    filters.append(("period_start_jalali", spreadsheet_datetime(report.period_start)))
    filters.append(("period_end", report.period_end))
    filters.append(("period_end_jalali", spreadsheet_datetime(report.period_end)))
    filters.append(("user_id", report.user_id))
    filters.append(("sales_product_id", report.sales_product_id))

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _write_money_columns(sheet, first_column, last_column):
    """Force money columns to text so a spreadsheet never re-rounds a Decimal."""
    if sheet.max_row <= 1:
        return
    for cells in sheet.iter_rows(min_row=2, min_col=first_column, max_col=last_column):
        for cell in cells:
            cell.number_format = "@"


def _new_workbook(title):
    workbook = Workbook()
    workbook.iso_dates = True
    workbook.properties.creator = "ForooshBin"
    sheet = workbook.active
    sheet.title = title
    sheet.freeze_panes = "A2"
    return workbook, sheet


def _finish(workbook, sheet):
    sheet.auto_filter.ref = sheet.dimensions
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


RECEIVABLES_HEADERS = (
    "customer_id", "customer_name", "invoice_count", "total_outstanding",
    "not_due", "days_1_30", "days_31_60", "days_61_90", "days_over_90",
)


def build_receivables_workbook(report):
    workbook, sheet = _new_workbook("receivables")
    sheet.append(RECEIVABLES_HEADERS)
    for row in report.results:
        sheet.append((
            row.customer_id,
            safe_spreadsheet_text(row.customer_name),
            row.invoice_count,
            format(row.total_outstanding, ".2f"),
            format(row.not_due, ".2f"),
            format(row.days_1_30, ".2f"),
            format(row.days_31_60, ".2f"),
            format(row.days_61_90, ".2f"),
            format(row.days_over_90, ".2f"),
        ))
    _write_money_columns(sheet, 4, 9)

    summary = workbook.create_sheet("summary")
    summary.append(("metric", "value"))
    summary.append(("as_of", spreadsheet_datetime(report.as_of)))
    summary.append(("total_outstanding", format(report.total_outstanding, ".2f")))
    for name, value in report.buckets.items():
        summary.append((name, format(value, ".2f")))
    _write_money_columns(summary, 2, 2)
    return _finish(workbook, sheet)


PROFIT_HEADERS = (
    "invoice_id", "number", "customer_id", "customer_name", "issued_at",
    "revenue", "cost", "profit", "margin_percent",
)


def build_profit_workbook(report):
    workbook, sheet = _new_workbook("profit")
    sheet.append(PROFIT_HEADERS)
    for row in report.results:
        sheet.append((
            row.invoice_id,
            safe_spreadsheet_text(row.number),
            row.customer_id,
            safe_spreadsheet_text(row.customer_name),
            spreadsheet_datetime(row.issued_at),
            format(row.revenue, ".2f"),
            format(row.cost, ".2f"),
            format(row.profit, ".2f"),
            format(row.margin_percent, ".2f"),
        ))
    _write_money_columns(sheet, 6, 9)

    summary = workbook.create_sheet("summary")
    summary.append(("metric", "value"))
    summary.append(("period_start", spreadsheet_datetime(report.period_start)))
    summary.append(("period_end", spreadsheet_datetime(report.period_end)))
    summary.append(("revenue", format(report.revenue, ".2f")))
    summary.append(("cost", format(report.cost, ".2f")))
    summary.append(("profit", format(report.profit, ".2f")))
    summary.append(("margin_percent", format(report.margin_percent, ".2f")))
    summary.append(("measured_invoice_count", report.measured_invoice_count))
    summary.append(("unmeasured_invoice_count", report.unmeasured_invoice_count))
    return _finish(workbook, sheet)


VALUATION_HEADERS = (
    "warehouse_id", "warehouse_name", "product_id", "product_sku", "product_name",
    "quantity", "average_cost", "stock_value",
)


def build_inventory_valuation_workbook(report):
    workbook, sheet = _new_workbook("stock-valuation")
    sheet.append(VALUATION_HEADERS)
    for row in report.results:
        sheet.append((
            row.warehouse_id,
            safe_spreadsheet_text(row.warehouse_name),
            row.product_id,
            safe_spreadsheet_text(row.product_sku),
            safe_spreadsheet_text(row.product_name),
            row.quantity,
            format(row.average_cost, ".2f"),
            format(row.stock_value, ".2f"),
        ))
    _write_money_columns(sheet, 7, 8)

    summary = workbook.create_sheet("summary")
    summary.append(("metric", "value"))
    summary.append(("as_of", spreadsheet_datetime(report.as_of)))
    summary.append(("total_quantity", report.total_quantity))
    summary.append(("total_value", format(report.total_value, ".2f")))
    return _finish(workbook, sheet)


USER_DIRECTORY_HEADERS = (
    "user_id", "username", "first_name", "last_name", "role",
    "workstream", "email", "phone", "is_active", "date_joined",
)


def build_user_directory_workbook(users):
    """The CRM user directory (requirement 1.9).

    No password, session key, or other credential material appears here: this is
    a directory, and an export is exactly the kind of file that leaves the
    building.
    """
    workbook, sheet = _new_workbook("users")
    sheet.append(USER_DIRECTORY_HEADERS)
    for user in users:
        sheet.append((
            user.pk,
            safe_spreadsheet_text(user.username),
            safe_spreadsheet_text(user.first_name),
            safe_spreadsheet_text(user.last_name),
            safe_spreadsheet_text(user.role),
            safe_spreadsheet_text(user.workstream),
            safe_spreadsheet_text(user.email),
            safe_spreadsheet_text(getattr(user, "phone", "") or ""),
            "yes" if user.is_active else "no",
            spreadsheet_datetime(user.date_joined),
        ))
    return _finish(workbook, sheet)


CUSTOMER_DIRECTORY_HEADERS = (
    "customer_id", "full_name", "primary_phone", "national_id", "email",
    "province", "city", "postal_code", "category", "is_active",
    "created_by", "created_at",
)


def build_customer_directory_workbook(customers):
    """The customer directory in the caller's scope (requirement 2.6)."""
    workbook, sheet = _new_workbook("customers")
    sheet.append(CUSTOMER_DIRECTORY_HEADERS)
    for customer in customers:
        primary = next(
            (phone for phone in customer.phones.all() if phone.is_primary and phone.is_active),
            None,
        )
        sheet.append((
            customer.pk,
            safe_spreadsheet_text(customer.full_name),
            safe_spreadsheet_text(primary.normalized_phone if primary else ""),
            safe_spreadsheet_text(customer.national_id),
            safe_spreadsheet_text(customer.email),
            safe_spreadsheet_text(customer.province),
            safe_spreadsheet_text(customer.city),
            safe_spreadsheet_text(customer.postal_code),
            safe_spreadsheet_text(customer.category),
            "yes" if customer.is_active else "no",
            safe_spreadsheet_text(customer.created_by.username if customer.created_by else ""),
            spreadsheet_datetime(customer.created_at),
        ))
    return _finish(workbook, sheet)


#: The product export and import share this header row exactly.
#:
#: That sharing is the whole design of the round trip: the operator exports,
#: writes on the file, and returns it, so the importer can map columns by name
#: instead of guessing at positions. `id` is present so an export stays useful
#: as a reference, and is ignored on the way back in — an import always creates.
PRODUCT_HEADERS = (
    "id",
    "sku",
    "name",
    "category_code",
    "brand",
    "unit",
    "current_price",
    "description",
    "is_active",
)

#: Persian labels the operator may type in the `unit` column, mapped to the
#: stored code. The stored codes are accepted too, so a returned export needs no
#: translation.
PRODUCT_UNIT_INPUT = {
    "جعبه": "box",
    "عدد": "piece",
    "کارتن": "carton",
    "کیلوگرم": "kilogram",
    "گرم": "gram",
    "box": "box",
    "piece": "piece",
    "carton": "carton",
    "kilogram": "kilogram",
    "gram": "gram",
}


def build_product_catalogue_workbook(products):
    """The product catalogue, in the exact shape the importer reads back."""
    workbook, sheet = _new_workbook("products")
    sheet.append(PRODUCT_HEADERS)
    for product in products:
        sheet.append((
            product.pk,
            safe_spreadsheet_text(product.sku),
            safe_spreadsheet_text(product.name),
            safe_spreadsheet_text(product.category.code if product.category else ""),
            safe_spreadsheet_text(product.brand),
            safe_spreadsheet_text(product.get_unit_display() if product.unit else ""),
            # Written as a plain number so the operator can edit it as one; the
            # panel adds grouping and the rial label when it displays it.
            product.current_price,
            safe_spreadsheet_text(product.description),
            "yes" if product.is_active else "no",
        ))
    return _finish(workbook, sheet)
