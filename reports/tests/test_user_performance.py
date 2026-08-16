from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from unittest import mock
from urllib.parse import urlencode

from django.core.cache import cache
from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from openpyxl import load_workbook
from rest_framework.test import APIClient

from accounts.models import User
from common.throttles import SensitiveRateThrottle
from reports.services import (
    InvalidReportPeriod,
    ReportAccessDenied,
    UserPerformanceReport,
    UserPerformanceRow,
    build_user_performance_report,
)
from reports.xlsx import (
    XLSX_CONTENT_TYPE,
    build_user_performance_workbook,
    safe_spreadsheet_text,
)
from sales.models import Customer, Lead, Product, Sale


class UserPerformanceReportTests(TestCase):
    report_url = "/api/v1/reports/user-performance/"
    detail_url = "/api/v1/reports/user-performance/details/"
    export_url = "/api/v1/exports/user-performance.xlsx"
    period_start = datetime(2026, 1, 1, tzinfo=UTC)
    period_end = datetime(2026, 2, 1, tzinfo=UTC)

    def setUp(self):
        self.agent = self._user("agent", User.Role.SALES_AGENT)
        self.manager = self._user("manager", User.Role.SALES_MANAGER)
        self.company_it = self._user("company-it", User.Role.COMPANY_IT)
        self.platform_admin = self._user("platform", User.Role.PLATFORM_ADMIN)
        self.formula_user = self._user("+former", User.Role.SALES_AGENT, is_active=False)

        self.product_a = Product.objects.create(
            sku="REPORT-A",
            name="Report A",
            current_price=Decimal("100.00"),
            created_by=self.manager,
            updated_by=self.manager,
        )
        self.product_b = Product.objects.create(
            sku="REPORT-B",
            name="Report B",
            current_price=Decimal("50.00"),
            created_by=self.manager,
            updated_by=self.manager,
        )

        self.agent_customer = self._customer(
            self.agent,
            "Agent customer",
            self.period_start,
        )
        self.manager_customer = self._customer(
            self.manager,
            "Manager customer",
            self.period_start + timedelta(days=5),
        )
        self.formula_customer = self._customer(
            self.formula_user,
            "Past customer",
            self.period_start + timedelta(days=6),
            is_active=False,
        )
        self._customer(
            self.agent,
            "End customer",
            self.period_end,
        )

        self._sale(
            self.agent,
            self.agent_customer,
            self.product_a,
            Decimal("100.00"),
            self.period_start,
        )
        self._sale(
            self.agent,
            self.agent_customer,
            self.product_b,
            Decimal("50.00"),
            self.period_end - timedelta(microseconds=1),
        )
        self._sale(
            self.agent,
            self.agent_customer,
            self.product_a,
            Decimal("900.00"),
            self.period_start + timedelta(days=10),
            status=Sale.Status.CANCELLED,
        )
        self._sale(
            self.formula_user,
            self.formula_customer,
            self.product_a,
            Decimal("30.00"),
            self.period_start + timedelta(days=11),
        )
        self._sale(
            self.manager,
            self.manager_customer,
            self.product_a,
            Decimal("20.00"),
            self.period_end,
        )
        Product.objects.filter(pk=self.product_b.pk).update(is_active=False)
        self.product_b.refresh_from_db()

    def _user(self, username, role, *, is_active=True):
        return User.objects.create_user(
            username=username,
            password="Long-Safe-Pass-741!",
            role=role,
            is_active=is_active,
        )

    def _customer(self, user, name, created_at, *, is_active=True):
        customer = Customer.objects.create(
            full_name=name,
            created_by=user,
            is_active=is_active,
        )
        Customer.objects.filter(pk=customer.pk).update(created_at=created_at)
        customer.refresh_from_db()
        return customer

    def _sale(self, user, customer, product, amount, sold_at, *, status=Sale.Status.CONFIRMED):
        lead = Lead.objects.create(customer=customer, created_by=user)
        return Sale.objects.create(
            lead=lead,
            customer=customer,
            sold_by=user,
            product=product,
            quantity=1,
            unit_price_snapshot=amount,
            total_amount=amount,
            status=status,
            sold_at=sold_at,
        )

    def _query(self, **overrides):
        query = {
            "period_start": "2026-01-01T00:00:00Z",
            "period_end": "2026-02-01T00:00:00Z",
        }
        query.update(overrides)
        return query

    def _get(self, actor, url=None, **query):
        client = APIClient()
        client.force_authenticate(actor)
        return client.get(url or self.report_url, self._query(**query))

    def test_agent_gets_only_exact_own_metrics(self):
        response = self._get(self.agent)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Cache-Control"], "private, no-store")
        self.assertEqual(response.data["period_start"], "2026-01-01T00:00:00Z")
        self.assertEqual(response.data["period_end"], "2026-02-01T00:00:00Z")
        self.assertEqual(
            response.data["results"],
            [
                {
                    "user_id": self.agent.pk,
                    "username": self.agent.username,
                    "customers_created_count": 1,
                    "sales_count": 2,
                    "sales_amount": "150.00",
                    "average_sale_amount": "75.00",
                }
            ],
        )
        self.assertEqual(
            response.data["summary"],
            {
                "customers_created_count": 1,
                "sales_count": 2,
                "sales_amount": "150.00",
                "average_sale_amount": "75.00",
            },
        )

    def test_product_filter_changes_only_confirmed_sale_metrics(self):
        response = self._get(self.agent, sales_product_id=self.product_a.pk)
        self.assertEqual(response.status_code, 200)
        row = response.data["results"][0]
        self.assertEqual(row["customers_created_count"], 1)
        self.assertEqual(row["sales_count"], 1)
        self.assertEqual(row["sales_amount"], "100.00")
        self.assertEqual(row["average_sale_amount"], "100.00")

        inactive_product = self._get(self.agent, sales_product_id=self.product_b.pk)
        self.assertEqual(inactive_product.status_code, 200)
        inactive_row = inactive_product.data["results"][0]
        self.assertEqual(inactive_row["sales_count"], 1)
        self.assertEqual(inactive_row["sales_amount"], "50.00")

        missing_product = self._get(self.agent, sales_product_id=999999)
        self.assertEqual(missing_product.status_code, 200)
        missing_row = missing_product.data["results"][0]
        self.assertEqual(missing_row["customers_created_count"], 1)
        self.assertEqual(missing_row["sales_count"], 0)
        self.assertEqual(missing_row["sales_amount"], "0.00")

        hidden_product = Product.objects.create(
            sku="REPORT-HIDDEN",
            name="Hidden historical product",
            current_price=Decimal("25.00"),
            is_active=False,
            created_by=self.manager,
            updated_by=self.manager,
        )
        hidden_product_response = self._get(
            self.agent,
            sales_product_id=hidden_product.pk,
        )
        self.assertEqual(hidden_product_response.status_code, 200)
        self.assertEqual(
            hidden_product_response.data["results"],
            missing_product.data["results"],
        )

    def test_elevated_roles_get_stable_company_rows_and_inactive_history(self):
        expected_ids = [
            self.agent.pk,
            self.manager.pk,
            self.company_it.pk,
            self.platform_admin.pk,
            self.formula_user.pk,
        ]
        for actor in (self.manager, self.company_it, self.platform_admin):
            with self.subTest(role=actor.role):
                response = self._get(actor)
                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    [row["user_id"] for row in response.data["results"]],
                    expected_ids,
                )
                inactive_row = next(
                    row for row in response.data["results"] if row["user_id"] == self.formula_user.pk
                )
                self.assertEqual(inactive_row["customers_created_count"], 1)
                self.assertEqual(inactive_row["sales_count"], 1)
                self.assertEqual(inactive_row["sales_amount"], "30.00")
                manager_row = next(
                    row for row in response.data["results"] if row["user_id"] == self.manager.pk
                )
                self.assertEqual(manager_row["sales_count"], 0)
                self.assertEqual(manager_row["average_sale_amount"], "0.00")

    def test_user_filter_obeys_scope_without_user_enumeration(self):
        selected = self._get(self.manager, user_id=self.formula_user.pk)
        self.assertEqual(selected.status_code, 200)
        self.assertEqual(
            [row["user_id"] for row in selected.data["results"]],
            [self.formula_user.pk],
        )

        own = self._get(self.agent, user_id=self.agent.pk)
        self.assertEqual(own.status_code, 200)

        hidden = self._get(self.agent, user_id=self.manager.pk)
        missing = self._get(self.agent, user_id=999999)
        hidden_export = self._get(self.agent, self.export_url, user_id=self.manager.pk)
        self.assertEqual(hidden.status_code, 400)
        self.assertEqual(missing.status_code, 400)
        self.assertEqual(hidden_export.status_code, 400)
        self.assertEqual(hidden.data["user_id"], missing.data["user_id"])
        self.assertEqual(hidden_export.data["user_id"], hidden.data["user_id"])
        self.assertEqual(hidden.data["error"]["code"], "validation_error")

    def test_detail_rows_match_aggregate_scope_and_product_filter(self):
        company_customers = self._get(
            self.manager,
            self.detail_url,
            metric="customers_created_count",
        )
        self.assertEqual(company_customers.status_code, 200)
        self.assertEqual(company_customers.data["count"], 3)
        self.assertEqual({row["record_type"] for row in company_customers.data["results"]}, {"customer"})

        agent_sales = self._get(
            self.agent,
            self.detail_url,
            metric="sales_count",
            user_id=self.agent.pk,
            sales_product_id=self.product_a.pk,
        )
        self.assertEqual(agent_sales.status_code, 200)
        self.assertEqual(agent_sales.data["count"], 1)
        self.assertEqual(agent_sales.data["results"][0]["amount"], "100.00")
        self.assertEqual(agent_sales.data["results"][0]["owner"], self.agent.username)
        self.assertEqual(agent_sales.data["results"][0]["detail_url"], f"/sales/{self.agent_customer.sales.get(status=Sale.Status.CONFIRMED, product=self.product_a).pk}/")

    def test_detail_user_id_is_fail_closed_for_agent(self):
        hidden = self._get(
            self.agent,
            self.detail_url,
            metric="sales_count",
            user_id=self.manager.pk,
        )
        missing = self._get(
            self.agent,
            self.detail_url,
            metric="sales_count",
            user_id=999999,
        )
        self.assertEqual(hidden.status_code, 400)
        self.assertEqual(missing.status_code, 400)
        self.assertEqual(hidden.data["user_id"], missing.data["user_id"])

    def test_query_contract_rejects_unknown_repeated_naive_and_bad_ranges(self):
        client = APIClient()
        client.force_authenticate(self.agent)

        unknown = client.get(self.report_url, self._query(extra="value"))
        self.assertEqual(unknown.status_code, 400)
        self.assertEqual([str(item) for item in unknown.data["extra"]], ["Unknown field."])

        repeated_query = urlencode(
            [
                ("period_start", "2026-01-01T00:00:00Z"),
                ("period_start", "2026-01-02T00:00:00Z"),
                ("period_end", "2026-02-01T00:00:00Z"),
            ]
        )
        repeated = client.get(f"{self.report_url}?{repeated_query}")
        self.assertEqual(repeated.status_code, 400)
        self.assertEqual(
            [str(item) for item in repeated.data["period_start"]],
            ["Query parameter must appear once."],
        )

        naive = client.get(
            self.report_url,
            self._query(period_start="2026-01-01T00:00:00"),
        )
        self.assertEqual(naive.status_code, 400)

        reversed_period = client.get(
            self.report_url,
            self._query(
                period_start="2026-02-01T00:00:00Z",
                period_end="2026-01-01T00:00:00Z",
            ),
        )
        self.assertEqual(reversed_period.status_code, 400)

        bad_product = self._get(self.agent, sales_product_id=0)
        self.assertEqual(bad_product.status_code, 400)

    def test_json_and_xlsx_use_same_filtered_result_and_formula_defense(self):
        query = {
            "user_id": self.formula_user.pk,
            "sales_product_id": self.product_a.pk,
        }
        json_response = self._get(self.manager, **query)
        export_response = self._get(self.manager, self.export_url, **query)
        self.assertEqual(json_response.status_code, 200)
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(export_response["Content-Type"], XLSX_CONTENT_TYPE)
        self.assertEqual(
            export_response["Content-Disposition"],
            'attachment; filename="forooshbin-user-performance.xlsx"',
        )
        self.assertEqual(export_response["Cache-Control"], "private, no-store")

        workbook = load_workbook(BytesIO(export_response.content), data_only=False)
        sheet = workbook["user-performance"]
        self.assertEqual(
            tuple(cell.value for cell in sheet[1]),
            (
                "user_id",
                "username",
                "customers_created_count",
                "sales_count",
                "sales_amount",
                "average_sale_amount",
            ),
        )
        json_row = json_response.data["results"][0]
        self.assertEqual(sheet["A2"].value, json_row["user_id"])
        self.assertEqual(sheet["B2"].value, "'+former")
        self.assertEqual(sheet["B2"].data_type, "s")
        self.assertEqual(sheet["C2"].value, json_row["customers_created_count"])
        self.assertEqual(sheet["D2"].value, json_row["sales_count"])
        self.assertEqual(sheet["E2"].value, json_row["sales_amount"])
        self.assertEqual(sheet["F2"].value, json_row["average_sale_amount"])
        self.assertEqual(sheet["E2"].data_type, "s")
        self.assertEqual(sheet["F2"].data_type, "s")

        summary = dict(workbook["summary"].iter_rows(min_row=2, values_only=True))
        json_summary = json_response.data["summary"]
        self.assertEqual(summary["customers_created_count"], json_summary["customers_created_count"])
        self.assertEqual(summary["sales_count"], json_summary["sales_count"])
        self.assertEqual(summary["sales_amount"], json_summary["sales_amount"])
        self.assertEqual(summary["average_sale_amount"], json_summary["average_sale_amount"])

        filters = dict(workbook["filters"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(filters["period_start"], json_response.data["period_start"])
        self.assertEqual(filters["period_end"], json_response.data["period_end"])
        self.assertEqual(filters["user_id"], json_response.data["user_id"])
        self.assertEqual(
            filters["sales_product_id"],
            json_response.data["sales_product_id"],
        )
        workbook.close()

    def test_export_accepts_xlsx_but_keeps_all_errors_json(self):
        client = APIClient()
        client.force_authenticate(self.manager)
        success = client.get(
            self.export_url,
            self._query(),
            HTTP_ACCEPT=XLSX_CONTENT_TYPE,
        )
        self.assertEqual(success.status_code, 200)
        self.assertEqual(success["Content-Type"], XLSX_CONTENT_TYPE)

        invalid = client.get(
            self.export_url,
            {"period_start": "2026-01-01T00:00:00Z"},
            HTTP_ACCEPT=XLSX_CONTENT_TYPE,
            HTTP_X_REQUEST_ID="xlsx-error-1",
        )
        self.assertEqual(invalid.status_code, 400)
        self.assertEqual(invalid["Content-Type"], "application/json")
        self.assertEqual(invalid.data["error"]["code"], "validation_error")
        self.assertEqual(invalid.data["error"]["request_id"], "xlsx-error-1")

        anonymous = APIClient().get(
            self.export_url,
            self._query(),
            HTTP_ACCEPT=XLSX_CONTENT_TYPE,
        )
        self.assertEqual(anonymous.status_code, 403)
        self.assertEqual(anonymous["Content-Type"], "application/json")
        self.assertEqual(anonymous.data["error"]["code"], "authentication_failed")

        unacceptable = client.get(
            self.export_url,
            self._query(),
            HTTP_ACCEPT="text/html",
        )
        self.assertEqual(unacceptable.status_code, 406)
        self.assertEqual(unacceptable["Content-Type"], "application/json")
        self.assertEqual(unacceptable.data["error"]["code"], "not_acceptable")

    @mock.patch.object(SensitiveRateThrottle, "get_rate", lambda self: "1/min")
    def test_export_xlsx_throttle_error_is_json(self):
        cache.clear()
        try:
            client = APIClient()
            client.force_authenticate(self.manager)
            first = client.get(
                self.export_url,
                self._query(),
                HTTP_ACCEPT=XLSX_CONTENT_TYPE,
            )
            self.assertEqual(first.status_code, 200)

            throttled = client.get(
                self.export_url,
                self._query(),
                HTTP_ACCEPT=XLSX_CONTENT_TYPE,
                HTTP_X_REQUEST_ID="xlsx-throttle-1",
            )
            self.assertEqual(throttled.status_code, 429)
            self.assertEqual(throttled["Content-Type"], "application/json")
            self.assertEqual(throttled.data["error"]["code"], "throttled")
            self.assertEqual(
                throttled.data["error"]["request_id"],
                "xlsx-throttle-1",
            )
        finally:
            cache.clear()

    def test_spreadsheet_text_guard_covers_formula_and_control_prefixes(self):
        unsafe_values = (
            "=SUM(A1:A2)",
            "+cmd",
            "-cmd",
            "@cmd",
            "  =SUM(A1:A2)",
            " \t=cmd",
            "\ttext",
            "\rtext",
            "\ntext",
        )
        for value in unsafe_values:
            with self.subTest(value=repr(value)):
                self.assertEqual(safe_spreadsheet_text(value), f"'{value}")
        self.assertEqual(safe_spreadsheet_text("safe text"), "safe text")

    def test_workbook_preserves_large_money_without_excel_rounding(self):
        report = UserPerformanceReport(
            period_start="2026-01-01T00:00:00Z",
            period_end="2026-02-01T00:00:00Z",
            user_id=1,
            sales_product_id=None,
            results=(
                UserPerformanceRow(
                    user_id=1,
                    username="money-proof",
                    customers_created_count=0,
                    sales_count=2,
                    sales_amount=Decimal("19999999999999999.98"),
                    average_sale_amount=Decimal("9999999999999999.99"),
                ),
            ),
        )
        workbook = load_workbook(
            BytesIO(build_user_performance_workbook(report)),
            data_only=False,
        )
        sheet = workbook["user-performance"]
        self.assertEqual(sheet["E2"].value, "19999999999999999.98")
        self.assertEqual(sheet["F2"].value, "9999999999999999.99")
        self.assertEqual(sheet["E2"].data_type, "s")
        self.assertEqual(sheet["F2"].data_type, "s")
        workbook.close()

    def test_report_query_count_stays_constant(self):
        with CaptureQueriesContext(connection) as captured:
            report = build_user_performance_report(
                actor=self.manager,
                period_start=self.period_start,
                period_end=self.period_end,
            )
        select_queries = [
            query
            for query in captured
            if query["sql"].lstrip().upper().startswith("SELECT")
        ]
        self.assertEqual(len(select_queries), 4)
        self.assertEqual(len(report.results), 5)

    def test_company_summary_uses_exact_totals_and_weighted_average(self):
        response = self._get(self.manager)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.data["summary"],
            {
                "customers_created_count": 3,
                "sales_count": 3,
                "sales_amount": "180.00",
                "average_sale_amount": "60.00",
            },
        )

    def test_service_scope_uses_fresh_locked_actor_state(self):
        stale_manager = self.manager
        User.objects.filter(pk=stale_manager.pk).update(role=User.Role.SALES_AGENT)

        report = build_user_performance_report(
            actor=stale_manager,
            period_start=self.period_start,
            period_end=self.period_end,
        )

        self.assertEqual(stale_manager.role, User.Role.SALES_MANAGER)
        self.assertEqual(
            [row.user_id for row in report.results],
            [stale_manager.pk],
        )
        User.objects.filter(pk=stale_manager.pk).update(is_active=False)
        with self.assertRaises(ReportAccessDenied):
            build_user_performance_report(
                actor=stale_manager,
                period_start=self.period_start,
                period_end=self.period_end,
            )

    def test_server_managed_user_is_not_enumerated_in_report_rows(self):
        server_user = User.objects.create_user(
            username="server-report-user",
            password="Server-Only-Pass-741!",
            role=User.Role.SALES_AGENT,
            is_staff=True,
        )
        response = self._get(self.manager)
        self.assertEqual(response.status_code, 200)
        self.assertNotIn(server_user.pk, [row["user_id"] for row in response.data["results"]])

    def test_service_rejects_bad_period_and_rounds_average_half_up(self):
        with self.assertRaises(InvalidReportPeriod):
            build_user_performance_report(
                actor=self.manager,
                period_start=self.period_end,
                period_end=self.period_start,
            )
        with self.assertRaises(InvalidReportPeriod):
            build_user_performance_report(
                actor=self.manager,
                period_start=datetime(2026, 1, 1),
                period_end=datetime(2026, 2, 1),
            )

        self._sale(
            self.company_it,
            self.manager_customer,
            self.product_a,
            Decimal("10.00"),
            self.period_start + timedelta(days=12),
        )
        self._sale(
            self.company_it,
            self.manager_customer,
            self.product_a,
            Decimal("10.01"),
            self.period_start + timedelta(days=13),
        )
        response = self._get(self.company_it, user_id=self.company_it.pk)
        self.assertEqual(response.status_code, 200)
        row = response.data["results"][0]
        self.assertEqual(row["sales_amount"], "20.01")
        self.assertEqual(row["average_sale_amount"], "10.01")

    def test_authentication_and_inactive_user_are_denied_for_both_formats(self):
        anonymous = APIClient()
        self.assertEqual(anonymous.get(self.report_url, self._query()).status_code, 403)
        self.assertEqual(anonymous.get(self.export_url, self._query()).status_code, 403)

        inactive = APIClient()
        inactive.force_authenticate(self.formula_user)
        self.assertEqual(inactive.get(self.report_url, self._query()).status_code, 403)
        self.assertEqual(inactive.get(self.export_url, self._query()).status_code, 403)

    def test_schema_documents_report_filters_and_binary_export(self):
        client = APIClient()
        client.force_authenticate(self.platform_admin)
        response = client.get(
            "/api/v1/schema/",
            HTTP_ACCEPT="application/vnd.oai.openapi+json",
        )
        self.assertEqual(response.status_code, 200)
        report_operation = response.data["paths"][self.report_url]["get"]
        export_operation = response.data["paths"][self.export_url]["get"]
        report_parameters = {item["name"]: item for item in report_operation["parameters"]}
        export_parameters = {item["name"]: item for item in export_operation["parameters"]}
        self.assertEqual(
            set(report_parameters),
            {"period_start", "period_end", "user_id", "sales_product_id"},
        )
        self.assertEqual(set(export_parameters), set(report_parameters))
        self.assertTrue(report_parameters["period_start"]["required"])
        self.assertTrue(report_parameters["period_end"]["required"])
        self.assertFalse(report_parameters["user_id"].get("required", False))
        self.assertIn(
            XLSX_CONTENT_TYPE,
            export_operation["responses"]["200"]["content"],
        )
        detail_parameters = {
            item["name"]
            for item in response.data["paths"][self.detail_url]["get"]["parameters"]
        }
        self.assertEqual(
            detail_parameters,
            {"period_start", "period_end", "user_id", "sales_product_id", "metric", "page"},
        )
